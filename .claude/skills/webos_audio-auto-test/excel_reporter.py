"""测试报告 Excel 写入器。

基于公司模板 QR-DQC-HT-202 音频性能测试报告(衍生)V5.0.xlsm，
把 main.py 跑出来的结果按 sheet / 行号映射填入对应单元格，并写入判定 (PASS/FAIL)。

模板结构要点（已实测确认）：
  Sheet '音频性能测试（中高音喇叭）'：
    每个通道 4 行，row_start 在 config.json 配置：
      row+0  标准输入功率  (K,L 填 W；M 填 PASS/FAIL)
      row+1  标准输入 THD+N (K,L 填 %；M 填 PASS/FAIL)
      row+2  最大输入功率
      row+3  最大输入 THD+N

  Sheet '中高音喇叭音量曲线（AV/ATV/HDMI）'：
    C 列=Volume (5..100 step 5)  从 row=3 开始共 20 行
    D 列=Left(W)  E 列=Right(W)  F 列=dB 公式自动计算

  Sheet '音频性能测试（耳机）'：
    每个通道 4 行，row_start 在 config.json 配置：
      row+0  标准输入 Vrms (K,L 填 mV；M 填 PASS/FAIL)
      row+1  标准输入 THD+N
      row+2  最大输入 Vrms
      row+3  最大输入 THD+N

  Sheet '耳机音量曲线'：
    同喇叭曲线格式，但 D/E 列是 Vrms (mV)
"""
from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional


def write_report(
    config: dict,
    speaker_perf: Optional[List[dict]] = None,
    speaker_curve: Optional[List[dict]] = None,
    hp_perf: Optional[List[dict]] = None,
    hp_curve: Optional[List[dict]] = None,
) -> str:
    """把测试结果写入 Excel 模板，返回输出文件路径。"""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("缺少 openpyxl，请先安装: pip install openpyxl") from exc

    # 模板路径：相对路径解析到 skill 目录，绝对路径直接使用
    skill_dir = Path(__file__).resolve().parent
    tpl_cfg = config.get("report_template", "template.xlsm")
    template_path = Path(tpl_cfg) if Path(tpl_cfg).is_absolute() else skill_dir / tpl_cfg
    if not template_path.exists():
        raise FileNotFoundError(f"未找到报告模板: {template_path}")

    out_dir = skill_dir / config.get("output_dir", "output/reports")
    ts = config.get("_session_ts", datetime.now().strftime("%Y%m%d_%H%M%S"))
    # 使用已有的 session 目录，或新建
    session_dir_str = config.get("_session_dir")
    if session_dir_str:
        session_dir = Path(session_dir_str)
    else:
        session_dir = out_dir / f"test_{ts}"
    session_dir.mkdir(parents=True, exist_ok=True)
    spec = _spec_tag(config)
    os_tag = config.get("os_type", "webOS")
    out_name = f"AudioTestReport_{os_tag}_{spec}_{ts}.xlsm"
    out_path = session_dir / out_name
    shutil.copy(template_path, out_path)

    wb = load_workbook(out_path, keep_vba=True)

    # 填写表头页（阻抗、额定功率等）
    _fill_header(wb, config)

    if speaker_perf is not None:
        _fill_speaker_performance(wb, config, speaker_perf)
    if speaker_curve is not None:
        _fill_speaker_curve(wb, config, speaker_curve)
    if hp_perf is not None:
        _fill_headphone_performance(wb, config, hp_perf)
    if hp_curve is not None:
        _fill_headphone_curve(wb, config, hp_curve)

    wb.save(out_path)
    return str(out_path)


def _spec_tag(config: dict) -> str:
    imp = str(config.get("impedance", "6R")).upper().replace(" ", "")
    pw = config.get("rated_power_w")
    try:
        pw_int = int(float(pw)) if pw is not None else None
    except (TypeError, ValueError):
        pw_int = None
    if pw_int is not None:
        return f"{imp}{pw_int}W"
    return imp


# ---------- 表头 ----------

def _fill_header(wb, config: dict) -> None:
    """填写表头页：中高音喇叭阻抗 (C5)、额定功率 (E5)。"""
    # 尝试多种可能的 sheet 名称
    header_sheet = None
    for name in wb.sheetnames:
        if "表头" in name or "header" in name.lower():
            header_sheet = name
            break
    if header_sheet is None:
        return
    ws = wb[header_sheet]

    imp = config.get("impedance", "")
    pw = config.get("rated_power_w")

    if imp:
        _safe_set_cell(ws, 5, 3, imp)       # C5: 中高音喇叭阻抗
    if pw is not None:
        _safe_set_cell(ws, 5, 5, f"{pw}W")  # E5: 中高音喇叭额定功率


# ---------- 喇叭性能 ----------

def _safe_set_cell(ws, row: int, column: int, value) -> None:
    """写入单元格，跳过合并单元格。"""
    from openpyxl.cell.cell import MergedCell
    cell = ws.cell(row=row, column=column)
    if isinstance(cell, MergedCell):
        return
    cell.value = value


def _fill_speaker_performance(wb, config: dict, results: List[dict]) -> None:
    sheet_name = config["speaker_performance"]["sheet"]
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    target_power = float(config["rated_power_w"])

    for ch in results:
        for case in ch.get("cases", []):
            if "error" in case:
                continue
            row = case["row"]
            l_val = case["L"]
            r_val = case["R"]
            if "thd" in case.get("kind", ""):
                l_val = l_val / 100.0
                r_val = r_val / 100.0

            # THD 行的 K:L 是合并单元格，只写 K 列（取 L/R 最大值）
            if "thd" in case.get("kind", ""):
                _safe_set_cell(ws, row, 11, max(l_val, r_val))
            else:
                _safe_set_cell(ws, row, 11, l_val)
                _safe_set_cell(ws, row, 12, r_val)

            if not case["passed"]:
                _safe_set_cell(ws, row, 14, "；".join(case.get("reasons", [])))


# ---------- 喇叭音量曲线 ----------

def _fill_speaker_curve(wb, config: dict, results: List[dict]) -> None:
    curve_cfg = config["speaker_curve"]
    for ch in results:
        if "error" in ch:
            continue
        sheet_name = curve_cfg["sheet"].get(ch["sheet"])
        if not sheet_name or sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        start_row = int(curve_cfg["data_start_row"])
        for i, pt in enumerate(ch.get("points", [])):
            row = start_row + i
            if "error" in pt:
                continue
            ws.cell(row=row, column=int(curve_cfg["left_column"]), value=pt["L"])
            ws.cell(row=row, column=int(curve_cfg["right_column"]), value=pt["R"])


# ---------- 耳机性能 ----------

def _fill_headphone_performance(wb, config: dict, results: List[dict]) -> None:
    sheet_name = config["headphone_performance"]["sheet"]
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]

    for ch in results:
        for case in ch.get("cases", []):
            if "error" in case:
                continue
            row = case["row"]
            l_val = case["L"]
            r_val = case["R"]
            if "thd" in case.get("kind", ""):
                l_val = l_val / 100.0
                r_val = r_val / 100.0

            # THD 行的 K:L 是合并单元格，只写 K 列（取 L/R 最大值）
            if "thd" in case.get("kind", ""):
                _safe_set_cell(ws, row, 11, max(l_val, r_val))
            else:
                _safe_set_cell(ws, row, 11, l_val)
                _safe_set_cell(ws, row, 12, r_val)
            if not case["passed"]:
                _safe_set_cell(ws, row, 14, "；".join(case.get("reasons", [])))


# ---------- 耳机音量曲线 ----------

def _fill_headphone_curve(wb, config: dict, results: List[dict]) -> None:
    curve_cfg = config["headphone_curve"]
    sheet_name = curve_cfg["sheet"]
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    start_row = int(curve_cfg["data_start_row"])
    for ch in results:
        if "error" in ch:
            continue
        for i, pt in enumerate(ch.get("points", [])):
            row = start_row + i
            if "error" in pt:
                continue
            ws.cell(row=row, column=int(curve_cfg["left_column"]), value=pt["L"])
            ws.cell(row=row, column=int(curve_cfg["right_column"]), value=pt["R"])
