"""音频功率报告生成器。

读取 output/reports/音频功率报告模版.xlsx，把 engine 跑出来的 standard / max 测量结果
按通道映射到对应行，生成带时间戳的报告文件。

标题 A1 的占位 `____欧____瓦` 由 power_spec 填充：
- auto-calibration / 常规场景：填 `<ohm>欧<watt>瓦`
- self-check 且用户仅提供 impedance（没有 rated_power）：只填 `<ohm>欧`，保留"瓦"前为空
"""
from __future__ import annotations

import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional


# 通道 → 模板行号映射（B/C 列对应 standard，E/F 列对应 max）
CHANNEL_ROW_MAP = {
    "AV":       3,
    "YPBPR":    4,
    "VGA":      5,
    "SCART":    6,
    "ATV-PAL":  7,
    "ATV-N":    8,
    "ATV-NTSC": 8,
    "DTV":      9,
    "DTV-DVB":  9,
    "DTV-ATSC": 9,
    "DTV-ISDB": 9,
    "HDMI":     10,
    "TYPEC":    11,
    "USB":      12,
}

HP_CHANNEL_ROW_MAP = {
    "AV":       17,
    "VGA":      18,
    "YPBPR":    18,
    "ATV-PAL":  19,
    "ATV-N":    20,
    "ATV-NTSC": 20,
    "HDMI":     21,
}

REPORTS_DIR = Path(__file__).resolve().parent / "output" / "reports"
TEMPLATE_FILENAME = "音频功率报告模版.xlsx"


def generate_report(
    measurements: List[Dict[str, Any]],
    output_path: Optional[str] = None,
    template_path: Optional[str] = None,
    power_spec: Optional[Dict[str, Any]] = None,
    mode: str = "auto_calibration",
    os_type: Optional[str] = None,
    output_type: str = "spk",
) -> str:
    """生成音频功率报告。

    measurements: engine._measure_channels 输出的 details["channels"]
    power_spec: {"impedance": "6R" or 6, "rated_power_w": 6.0}
                self-check 模式下 rated_power_w 可以缺省
    mode: "auto_calibration" | "self_check"
          self_check 且 power_spec 无 rated_power_w → 标题只填欧不填瓦
    os_type: 平台名称（"webOS" / "tizen" / "android"），写入文件名

    返回：实际写入的报告路径

    文件名规则（output_path 未指定时）：
        AudioCalibration_<os>_<NRMW>_<YYYYMMDD_HHMMSS>.xlsx          # auto-calibration
        AudioSelfCheck_<os>[_<NRMW>]_<YYYYMMDD_HHMMSS>.xlsx          # self-check
        其中 NRMW 形如 6R6W；self-check 缺 rated_power 时省略 _NRMW
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("缺少 openpyxl，请先安装: pip install openpyxl") from exc

    template = Path(template_path) if template_path else REPORTS_DIR / TEMPLATE_FILENAME
    if not template.exists():
        raise FileNotFoundError(f"未找到报告模版: {template}")

    if output_path:
        out = Path(output_path)
    else:
        out = REPORTS_DIR / _build_report_filename(power_spec or {}, mode, os_type)

    out.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(template, out)

    wb = load_workbook(out)
    ws = wb[wb.sheetnames[0]]

    # --- 1. 更新标题 A1 的 ____欧____瓦 占位 ---
    _fill_title(ws, power_spec or {}, mode)

    # --- 2. 写入每通道测量值 ---
    aggregated: Dict[str, Dict[str, Dict[str, Any]]] = {}
    for item in measurements:
        channel = str(item.get("channel", ""))
        test_case = item.get("test_case", {}) or {}
        kind = str(test_case.get("kind", "")).lower()
        if not channel or kind not in ("standard", "max"):
            continue
        aggregated.setdefault(channel, {})[kind] = item

    for channel, by_kind in aggregated.items():
        row_map = HP_CHANNEL_ROW_MAP if output_type == "hp" else CHANNEL_ROW_MAP
        row = row_map.get(channel)
        if not row:
            continue
        std_item = by_kind.get("standard")
        max_item = by_kind.get("max")

        if std_item:
            if output_type == "hp":
                avg_mv, max_thd = _avg_vrms_mv_and_max_thd(std_item)
                ws.cell(row=row, column=2, value=round(avg_mv, 2))
            else:
                avg_w, max_thd = _avg_power_and_max_thd(std_item)
                ws.cell(row=row, column=2, value=round(avg_w, 3))
            ws.cell(row=row, column=3, value=round(max_thd, 3))
        if max_item and output_type != "hp":
            avg_w, max_thd = _avg_power_and_max_thd(max_item)
            ws.cell(row=row, column=5, value=round(avg_w, 3))
            ws.cell(row=row, column=6, value=round(max_thd, 3))

    wb.save(out)
    return str(out)


def _fill_title(ws, power_spec: Dict[str, Any], mode: str) -> None:
    """把标题里的 `____欧____瓦` 替换为实际值。

    - 有阻抗 + 有功率 → `6欧6瓦`
    - self_check 且只有阻抗 → `6欧____瓦`（保留瓦占位）
    - 完全没有 → 不动
    """
    title_cell = ws.cell(row=1, column=1)
    title = title_cell.value
    if not isinstance(title, str):
        return

    ohm = _parse_ohm(power_spec.get("impedance"))
    watt = power_spec.get("rated_power_w")

    if ohm is None and watt is None:
        return

    is_self_check = str(mode).lower() in ("self_check", "selfcheck", "measure_only")
    ohm_text = _format_number(ohm) if ohm is not None else None
    watt_text = _format_number(float(watt)) if watt is not None else None

    # 标题里的占位：连续 4 个下划线
    placeholder = "____"

    def replace_first(text: str, old: str, new: str) -> str:
        idx = text.find(old)
        if idx < 0:
            return text
        return text[:idx] + new + text[idx + len(old):]

    new_title = title
    if ohm_text is not None:
        new_title = replace_first(new_title, placeholder, ohm_text)
    if watt_text is not None:
        new_title = replace_first(new_title, placeholder, watt_text)
    elif not is_self_check:
        # 非 self-check 场景缺功率，保留占位即可，不强制写
        pass

    title_cell.value = new_title


def _parse_ohm(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().upper().replace("Ω", "").replace("欧", "")
    if text.endswith("R"):
        text = text[:-1]
    match = re.search(r"[-+]?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group())
    except ValueError:
        return None


def _format_number(value: float) -> str:
    if value is None:
        return ""
    if float(value).is_integer():
        return str(int(value))
    return f"{value:g}"


def _avg_power_and_max_thd(item: Dict[str, Any]) -> tuple[float, float]:
    summary = item.get("summary") or {}
    if "avg_power_w" in summary and "max_thd_n_percent" in summary:
        return float(summary["avg_power_w"]), float(summary["max_thd_n_percent"])

    measurement = item.get("measurement") or {}
    powers = [float(v.get("power", 0.0)) for v in measurement.values()]
    thds = [float(v.get("thd_n", 0.0)) for v in measurement.values()]
    avg = sum(powers) / len(powers) if powers else 0.0
    max_thd = max(thds) if thds else 0.0
    return avg, max_thd


def _avg_vrms_mv_and_max_thd(item: Dict[str, Any]) -> tuple[float, float]:
    summary = item.get("summary") or {}
    if "avg_vrms_mv" in summary and "max_thd_n_percent" in summary:
        return float(summary["avg_vrms_mv"]), float(summary["max_thd_n_percent"])

    measurement = item.get("measurement") or {}
    vrms_mvs = [float(v.get("vrms_mv", v.get("power", 0.0) * 1000)) for v in measurement.values()]
    thds = [float(v.get("thd_n", 0.0)) for v in measurement.values()]
    avg = sum(vrms_mvs) / len(vrms_mvs) if vrms_mvs else 0.0
    max_thd = max(thds) if thds else 0.0
    return avg, max_thd


def _build_report_filename(power_spec: Dict[str, Any], mode: str, os_type: Optional[str]) -> str:
    """生成报告文件名。

    auto_calibration: AudioCalibration_<os>_<NRMW>_<ts>.xlsx   (阻抗+功率必带)
    self_check:       AudioSelfCheck_<os>[_<NRMW>]_<ts>.xlsx   (无功率则省略)
    """
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    is_self_check = str(mode).lower() in ("self_check", "selfcheck", "measure_only")
    prefix = "AudioSelfCheck" if is_self_check else "AudioCalibration"

    os_tag = _sanitize(os_type) if os_type else ""
    parts = [prefix]
    if os_tag:
        parts.append(os_tag)

    spec_tag = _build_spec_tag(power_spec, is_self_check)
    if spec_tag:
        parts.append(spec_tag)

    parts.append(ts)
    return "_".join(parts) + ".xlsx"


def _build_spec_tag(power_spec: Dict[str, Any], is_self_check: bool) -> str:
    ohm = _parse_ohm(power_spec.get("impedance"))
    watt = power_spec.get("rated_power_w")

    ohm_text = _format_number(ohm) if ohm is not None else ""
    watt_text = _format_number(float(watt)) if watt is not None else ""

    if is_self_check:
        if ohm_text and watt_text:
            return f"{ohm_text}R{watt_text}W"
        if ohm_text:
            return f"{ohm_text}R"
        return ""
    # auto-calibration: 要求阻抗 + 功率都有
    if ohm_text and watt_text:
        return f"{ohm_text}R{watt_text}W"
    if ohm_text:
        return f"{ohm_text}R"
    if watt_text:
        return f"{watt_text}W"
    return ""


def _sanitize(text: str) -> str:
    """清理文件名里不允许的字符。"""
    return re.sub(r"[^A-Za-z0-9._-]", "", str(text))
